"""Streamlit dashboard for tc_pipeline.

Deploy: point Streamlit Community Cloud at this file (main file path =
app.py). No credentials needed — this only processes an uploaded CSV in
memory and hands back a workbook.
"""

from __future__ import annotations

import io
import tempfile
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from tc_pipeline import ORDER_ID_COLUMN_LETTER_DEFAULT, run_pipeline

st.set_page_config(page_title="TC Report Processor", page_icon="📊", layout="wide")

st.title("📊 TC Report Processor")
st.caption(
    "Upload a TC Report CSV to generate a 3-sheet workbook: Raw Data, "
    "Final Report, Pivot Summary."
)

with st.sidebar:
    st.header("Options")
    order_id_column = st.text_input(
        "order_id column letter (for dedup only)",
        value=ORDER_ID_COLUMN_LETTER_DEFAULT,
        help="The spec dedups on order_id but never says which column "
        "that is (it isn't one of the retained Final Report columns). "
        "Defaults to Column A — change if your file differs.",
    )
    use_custom_today = st.checkbox(
        "Override 'today' for Pivot Summary coloring",
        value=False,
        help="Controls which date column is colored green vs red. Leave "
        "unchecked to use the live system date.",
    )
    custom_today = None
    if use_custom_today:
        custom_today = st.date_input("Today's date")

uploaded_file = st.file_uploader("TC Report CSV", type=["csv"])

if uploaded_file is None:
    st.info("Upload a CSV to get started.")
    st.stop()

run_clicked = st.button("▶ Process report", type="primary")

if run_clicked:
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "input.csv"
        input_path.write_bytes(uploaded_file.getvalue())
        output_path = Path(tmp) / "output.xlsx"

        try:
            with st.spinner("Processing..."):
                summary = run_pipeline(
                    input_csv=str(input_path),
                    output_xlsx=str(output_path),
                    order_id_column_letter=order_id_column or ORDER_ID_COLUMN_LETTER_DEFAULT,
                    today=custom_today,
                )
        except Exception as e:  # noqa: BLE001
            st.error(f"Processing failed: {e}")
            st.stop()

        output_bytes = output_path.read_bytes()

    st.success("Done.")

    if summary["column_warnings"]:
        for msg in summary["column_warnings"]:
            st.warning(f"⚠ {msg}")

    if summary["after_status_filter"] == 0:
        st.error(
            "0 rows survived the order-status filter (Step 2.1) out of "
            f"{summary['input_rows']} input rows. This almost always means "
            "either the wrong column was read as order_item_status, or "
            "your file's status text doesn't exactly match "
            "'New' / 'ACCEPTED/PICKED' / 'READY TO SHIP'."
        )
        st.write(
            f"**Column read as order_item_status:** "
            f"`{summary['order_item_status_column_resolved']}`"
        )
        if summary["order_item_status_value_counts"]:
            st.write("**Actual values found in that column, with counts:**")
            counts_df = pd.DataFrame(
                summary["order_item_status_value_counts"].items(),
                columns=["Value found in file", "Count"],
            )
            st.dataframe(counts_df, use_container_width=True, hide_index=True)
            st.caption(
                "If the column above isn't order_item_status, your file's "
                "columns don't line up with the letter positions this "
                "pipeline assumes (B, G, J, AP, AQ, AS, AV, BD, BI, BO, BP, "
                "BR) — check for extra/missing columns upstream. If it IS "
                "the right column but none of these values match "
                "'New' / 'ACCEPTED/PICKED' / 'READY TO SHIP', the status "
                "text in your file uses different wording/formatting than "
                "the spec's exact strings."
            )

    m1, m2, m3 = st.columns(3)
    m1.metric("Input rows", summary["input_rows"])
    m2.metric("After filtering", summary["after_erp_filter"])
    m3.metric("Final Report rows", summary["final_report_rows"])

    st.download_button(
        "⬇ Download workbook",
        data=output_bytes,
        file_name=f"TC_Report_Processed_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.subheader("Preview")
    wb = load_workbook(io.BytesIO(output_bytes))
    tabs = st.tabs(wb.sheetnames)
    for tab, sheet_name in zip(tabs, wb.sheetnames):
        with tab:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                st.write("(empty)")
                continue
            preview_df = pd.DataFrame(rows[1:], columns=rows[0])
            st.dataframe(preview_df.head(200), use_container_width=True, hide_index=True)
            if len(preview_df) > 200:
                st.caption(f"Showing first 200 of {len(preview_df)} rows.")
            if sheet_name == "Pivot Summary":
                st.caption(
                    "Color-coding (red = past dates, green = today) is visible "
                    "in the downloaded .xlsx, not in this plain preview table."
                )

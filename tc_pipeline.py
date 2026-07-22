"""tc_pipeline.py — everything needed to process a TC Report CSV into a
3-sheet workbook (Raw Data, Final Report, Pivot Summary), per the
"Prepare Pending Order Reports, Generate Final Report, Pivot Summary"
spec (this version drops the New Status Alert sheet and changes the
Final Report's retained columns).

Deliberately a single flat file, not a package/subfolder — a prior
version of this repo used a subpackage and it repeatedly failed to fully
upload to GitHub, causing ModuleNotFoundError on Streamlit Cloud. A lone
file next to app.py can't be partially uploaded the way a folder can.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pandas.api.types import is_string_dtype

logger = logging.getLogger("tc_report_processor")

# ===========================================================================
# Column resolution — everything used by this pipeline is identified by
# spreadsheet letter in the spec, so all of it is resolved by position.
# ===========================================================================

LETTER_COLUMNS: dict[str, str] = {
    "B": "order_number",
    "G": "payment_status",
    "J": "order_item_status",
    "AP": "courier_name",
    "AQ": "tracking_number",
    "AS": "ordered_date",
    "AV": "accepted_date",
    "BD": "nickname",
    "BI": "payment_methods",
    "BO": "time_shippinglabel_printed",
    "BP": "erp_reference_id",
    "BR": "time_order_paid",
}

# order_id is used only internally for dedup (Part 4: "Remove duplicate
# records using order_id") — it is NOT one of the spec's retained Final
# Report columns, and the spec doesn't say which column it is. Assuming
# Column A, the conventional primary-key position in these TC Report
# exports; override via --order-id-column if your file differs.
ORDER_ID_COLUMN_LETTER_DEFAULT = "A"

# Part 4: Final Report column order, exactly as specified (letter order
# in the spec, which is NOT the same as the sheet's original column
# order — e.g. order_item_status/J comes before payment_status/G here).
FINAL_REPORT_COLUMNS: list[str] = [
    "order_number",
    "order_item_status",
    "payment_status",
    "courier_name",
    "tracking_number",
    "ordered_date",
    "accepted_date",
    "nickname",
    "payment_methods",
    "time_shippinglabel_printed",
    "erp_reference_id",
    "time_order_paid",
]

# Step 2.1: order_item_status whitelist. The spec gives both a "remove
# these" list and a "retain only these" list; they're complementary for
# every status named. Implemented literally as a whitelist per "Retain
# only records with the following statuses" — anything not listed
# (including unexpected/typo values) is dropped.
ALLOWED_ORDER_STATUSES = {"new", "accepted/picked", "ready to ship"}

# Part 5: display label for the pivot's row dimension. The spec instructs
# "rows = Nickname (BD)" but the reference example's row header reads
# "Marketplace Channel" with values like "lazada-Hiruscar" — i.e. the
# nickname field's *content* already encodes the marketplace/channel, and
# the sheet just labels that column "Marketplace Channel" for readability.
# This pipeline pivots on nickname's actual values but writes this label
# as the header, matching the example.
PIVOT_ROW_LABEL = "Marketplace Channel"

PAST_DATE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TODAY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def col_letter_to_index(letter: str) -> int:
    """'A' -> 0, 'B' -> 1, ..., 'Z' -> 25, 'AA' -> 26, 'BP' -> 67, etc."""
    letter = letter.strip().upper()
    if not re.fullmatch(r"[A-Z]+", letter):
        raise ValueError(f"Invalid column letter: {letter!r}")
    index = 0
    for ch in letter:
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index - 1


def _normalize(name: str) -> str:
    return re.sub(r"[\s_]+", "", name.strip().lower())


def resolve_letter_columns(
    headers: list[str], letter_map: dict[str, str]
) -> dict[str, str]:
    """Returns {logical_name: actual_header_text}, resolved by position.
    Logs a warning (doesn't raise) if the header found there doesn't
    resemble the expected name."""
    resolved: dict[str, str] = {}
    for letter, expected_name in letter_map.items():
        idx = col_letter_to_index(letter)
        if idx >= len(headers):
            raise ValueError(
                f"Expected column {letter} ({expected_name}) at position {idx}, "
                f"but the file only has {len(headers)} columns. Check the file "
                f"matches the TC Report template this pipeline was built for."
            )
        actual_header = headers[idx]
        if _normalize(actual_header) != _normalize(expected_name):
            logger.warning(
                "Column %s: expected header resembling '%s' but found '%s'. "
                "Proceeding with the actual header text found there — verify "
                "this is really the right column before trusting the output.",
                letter,
                expected_name,
                actual_header,
            )
        resolved[expected_name] = actual_header
    return resolved


# ===========================================================================
# Loading
# ===========================================================================


def load_tc_report(
    csv_path: str, order_id_letter: str = ORDER_ID_COLUMN_LETTER_DEFAULT
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, str], str]:
    """Returns (raw_df, working_df, resolved_columns, order_id_header).
    raw_df: exact copy as read — becomes the untouched "Raw Data" sheet.
    working_df: a separate copy the rest of the pipeline mutates freely."""
    header_df = pd.read_csv(csv_path, nrows=0)
    headers = list(header_df.columns)

    resolved = resolve_letter_columns(headers, LETTER_COLUMNS)

    order_id_idx = col_letter_to_index(order_id_letter)
    if order_id_idx >= len(headers):
        raise ValueError(
            f"order_id column {order_id_letter} is out of range for a file "
            f"with {len(headers)} columns."
        )
    order_id_header = headers[order_id_idx]

    order_number_header = resolved["order_number"]
    erp_header = resolved["erp_reference_id"]

    raw_df = pd.read_csv(
        csv_path,
        dtype={order_number_header: str, erp_header: str},
        keep_default_na=True,
    )
    working_df = raw_df.copy(deep=True)
    return raw_df, working_df, resolved, order_id_header


# ===========================================================================
# Cleaning — Step 2: trim, order_number as text (handled at load).
# ===========================================================================


def trim_all_string_fields(df: pd.DataFrame) -> pd.DataFrame:
    """Handles both legacy `object` dtype and pandas's `StringDtype`
    (default for text columns as of pandas 3.0)."""
    df = df.copy()
    for col in df.columns:
        if is_string_dtype(df[col]):
            df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)
    return df


def is_blank(value) -> bool:
    if value is None:
        return True
    if value is pd.NA:
        return True
    try:
        if isinstance(value, float) and pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


# ===========================================================================
# Step 2.1 — order_item_status whitelist.
# ===========================================================================


def filter_by_order_status(df: pd.DataFrame, order_item_status_column: str) -> pd.DataFrame:
    status = df[order_item_status_column].astype(str).str.strip().str.casefold()
    mask = status.isin(ALLOWED_ORDER_STATUSES)
    return df[mask].copy()


# ===========================================================================
# Part 2 — ERP-blank filtering (New/Pending/non-COD removal).
# ===========================================================================


def filter_for_erp_investigation(
    df: pd.DataFrame,
    erp_column: str,
    order_item_status_column: str,
    payment_status_column: str,
    payment_methods_column: str,
) -> pd.DataFrame:
    """Only records where erp_reference_id is blank are evaluated (and
    thus survive at all). Of those, drop rows where order_item_status ==
    'New' AND payment_status == 'Pending' AND payment_methods != 'COD'.
    Everything else in the blank-erp subset survives."""
    blank_mask = df[erp_column].apply(is_blank)
    working = df[blank_mask].copy()

    status = working[order_item_status_column].astype(str).str.strip().str.casefold()
    payment_status = working[payment_status_column].astype(str).str.strip().str.casefold()
    payment_method = working[payment_methods_column].astype(str).str.strip().str.casefold()

    remove_mask = (status == "new") & (payment_status == "pending") & (payment_method != "cod")
    return working[~remove_mask].copy()


# ===========================================================================
# Part 4 — Final Report: select columns, THEN dedup by order_id.
# ===========================================================================


def build_final_report(
    df: pd.DataFrame,
    resolved_columns: dict[str, str],
    order_id_column: str,
) -> pd.DataFrame:
    if order_id_column not in df.columns:
        raise KeyError(
            f"order_id column '{order_id_column}' not found — needed for "
            f"dedup (Part 4) even though it isn't a retained output column."
        )

    columns_to_keep: list[str] = []
    for logical_name in FINAL_REPORT_COLUMNS:
        actual_header = resolved_columns.get(logical_name)
        if actual_header is None or actual_header not in df.columns:
            raise KeyError(
                f"Final Report column '{logical_name}' not resolved to an "
                f"actual header in the input file."
            )
        columns_to_keep.append(actual_header)

    # Dedup on order_id BEFORE dropping it from the retained columns (the
    # spec deletes "all remaining columns" after extracting the listed
    # ones, and order_id isn't one of them).
    deduped = df.drop_duplicates(subset=[order_id_column], keep="first")
    return deduped[columns_to_keep].copy()


# ===========================================================================
# Part 5 — Pivot Summary: nickname (displayed as "Marketplace Channel")
# x ordered_date, count of orders, with grand totals + red/green coloring.
# ===========================================================================


def build_pivot_summary(
    df: pd.DataFrame,
    nickname_column: str,
    ordered_date_column: str,
    count_column: str,
) -> pd.DataFrame:
    """Returns a pivot DataFrame whose date columns are real `date`
    objects (not pre-formatted strings) — formatting and color-coding
    happen at write time, in _write_pivot, where each column's real date
    is compared against "today" for the red/green rule."""
    working = df.copy()
    working["_ordered_date_only"] = pd.to_datetime(
        working[ordered_date_column], errors="coerce"
    ).dt.date

    pivot = pd.pivot_table(
        working,
        index=nickname_column,
        columns="_ordered_date_only",
        values=count_column,
        aggfunc="count",
        fill_value=0,
        margins=True,
        margins_name="Grand Total",
    )
    return pivot


# ===========================================================================
# Workbook writer — Part 6: Raw Data, Final Report, Pivot Summary.
# ===========================================================================

HEADER_FONT = Font(bold=True)
TEXT_FORMAT = "@"
TARGET_SHEET_ORDER = ["Raw Data", "Final Report", "Pivot Summary"]


def _write_df(ws: Worksheet, df: pd.DataFrame, text_columns: set[str] | None = None) -> None:
    text_columns = text_columns or set()
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT

    for row in df.itertuples(index=False):
        ws.append(list(row))

    header_to_idx = {col: i + 1 for i, col in enumerate(df.columns)}
    for col_name in text_columns:
        idx = header_to_idx.get(col_name)
        if idx is None:
            continue
        letter = get_column_letter(idx)
        for row_idx in range(2, ws.max_row + 1):
            ws[f"{letter}{row_idx}"].number_format = TEXT_FORMAT

    _autofit(ws, df)


def _autofit(ws: Worksheet, df: pd.DataFrame, max_width: int = 40) -> None:
    for i, col in enumerate(df.columns, start=1):
        sample = [str(col)] + [str(v) for v in df[col].astype(str).head(200)]
        width = min(max(len(s) for s in sample) + 2, max_width)
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_pivot(ws: Worksheet, pivot_df: pd.DataFrame, today: date) -> None:
    header = [PIVOT_ROW_LABEL] + [
        c.strftime("%d-%b") if isinstance(c, date) else str(c) for c in pivot_df.columns
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = HEADER_FONT

    for idx_value, row in pivot_df.iterrows():
        ws.append([idx_value] + list(row.values))

    grand_total_row_idx = next(
        (r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Grand Total"),
        None,
    )

    # Color each real date column: red if older than today, green if
    # today, no fill for future dates (not specified in the spec).
    # Applied to the header cell and every data cell in that column,
    # excluding the Grand Total row (kept bold/uncolored for contrast).
    for col_idx, col_value in enumerate(pivot_df.columns, start=2):
        if not isinstance(col_value, date):
            continue  # the "Grand Total" margin column
        if col_value < today:
            fill = PAST_DATE_FILL
        elif col_value == today:
            fill = TODAY_FILL
        else:
            continue  # future date — no fill specified by the spec
        for row_idx in range(1, ws.max_row + 1):
            if row_idx == grand_total_row_idx:
                continue
            ws.cell(row=row_idx, column=col_idx).fill = fill

    grand_total_col_idx = len(header)  # "Grand Total" margin is always last
    for r in range(1, ws.max_row + 1):
        ws.cell(row=r, column=grand_total_col_idx).font = HEADER_FONT
    if grand_total_row_idx:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=grand_total_row_idx, column=c).font = HEADER_FONT

    _autofit(ws, pd.DataFrame(columns=header))


def write_workbook(
    output_path: str,
    raw_df: pd.DataFrame,
    final_report_df: pd.DataFrame,
    pivot_df: pd.DataFrame,
    order_number_header: str,
    today: date,
) -> None:
    """If `output_path` already exists, loads it and replaces only the 3
    target sheets — any other sheet already in that workbook is left
    alone. Otherwise creates a fresh workbook."""
    existing = Path(output_path)
    if existing.exists():
        wb = load_workbook(output_path)
        for name in TARGET_SHEET_ORDER:
            if name in wb.sheetnames:
                del wb[name]
    else:
        wb = Workbook()
        wb.remove(wb.active)

    ws_raw = wb.create_sheet("Raw Data")
    _write_df(ws_raw, raw_df, text_columns={order_number_header})

    ws_final = wb.create_sheet("Final Report")
    _write_df(ws_final, final_report_df)

    ws_pivot = wb.create_sheet("Pivot Summary")
    _write_pivot(ws_pivot, pivot_df, today=today)

    other_sheets = [s for s in wb.sheetnames if s not in TARGET_SHEET_ORDER]
    wb._sheets = [wb[name] for name in TARGET_SHEET_ORDER] + [wb[name] for name in other_sheets]

    wb.save(output_path)


# ===========================================================================
# Orchestration
# ===========================================================================


def run_pipeline(
    input_csv: str,
    output_xlsx: str,
    order_id_column_letter: str = ORDER_ID_COLUMN_LETTER_DEFAULT,
    today: date | None = None,
) -> dict:
    today = today or datetime.now().date()

    raw_df, working_df, resolved, order_id_col = load_tc_report(
        input_csv, order_id_letter=order_id_column_letter
    )

    order_number_col = resolved["order_number"]
    payment_status_col = resolved["payment_status"]
    order_item_status_col = resolved["order_item_status"]
    payment_methods_col = resolved["payment_methods"]
    erp_col = resolved["erp_reference_id"]
    nickname_col = resolved["nickname"]
    ordered_date_col = resolved["ordered_date"]

    working_df = trim_all_string_fields(working_df)

    status_filtered_df = filter_by_order_status(working_df, order_item_status_col)

    filtered_df = filter_for_erp_investigation(
        status_filtered_df,
        erp_column=erp_col,
        order_item_status_column=order_item_status_col,
        payment_status_column=payment_status_col,
        payment_methods_column=payment_methods_col,
    )

    final_df = build_final_report(filtered_df, resolved_columns=resolved, order_id_column=order_id_col)

    pivot_df = build_pivot_summary(
        final_df,
        nickname_column=nickname_col,
        ordered_date_column=ordered_date_col,
        count_column=order_number_col,
    )

    write_workbook(
        output_xlsx,
        raw_df=raw_df,
        final_report_df=final_df,
        pivot_df=pivot_df,
        order_number_header=order_number_col,
        today=today,
    )

    return {
        "input_rows": len(raw_df),
        "after_status_filter": len(status_filtered_df),
        "after_erp_filter": len(filtered_df),
        "final_report_rows": len(final_df),
        "output_path": output_xlsx,
    }
